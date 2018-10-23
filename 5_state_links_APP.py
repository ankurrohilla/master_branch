import time
import openpyxl
from bs4 import BeautifulSoup
from selenium.common.exceptions import TimeoutException
import requests
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
options = webdriver.ChromeOptions()
options.add_argument("headless")
start = time.time()
counter = 0
def Clear_previous_data():
    file_name = "Data for 50 State.xlsx"
    wb = openpyxl.load_workbook(file_name)
    ws1 = wb.worksheets[0]
    for row in ws1['B2:L2345']:
        for cell in row:
            cell.value =None
    for ws in wb:
        for row in ws[2:ws.max_row]:
            if not ws.title == "Keywords":
                for cell in row:
                    cell.value = None
    wb.save(file_name)
    return Clear_previous_data
if __name__ == '__main__':
    Clear_previous_data = Clear_previous_data()
def Alabma_Trademark():
    file_name = "Data for 50 State.xlsx"
    wb = openpyxl.load_workbook(file_name)
    ws1 = wb.worksheets[0]
    N = range(2, ws1.max_row+1)
    j = 2
    J = 2
    for rownumber in N:
        keywords = ws1.cell(rownumber, 1).value
        if keywords is not None:
              driver = webdriver.Chrome(chrome_options=options)
              driver.get('http://arc-sos.state.al.us/CGI/TMMARK.MBR/INPUT')
              wait = WebDriverWait(driver, 10)
              element=wait.until(EC.visibility_of_any_elements_located((By.TAG_NAME, "input")))
              driver.find_element_by_xpath('//*[@id="block-sos-content"]/div/div/div[1]/form/div[1]/input').send_keys(keywords)
              driver.find_element_by_xpath('//*[@id="block-sos-content"]/div/div/div[1]/form/div[4]/input').click()
              element1 = wait.until(EC.visibility_of_any_elements_located((By.TAG_NAME, "input")))
              frame = driver.find_elements_by_tag_name('tbody')
              if not frame:
                  ws1.cell(rownumber, 3).value = 0
              else:
                  def scrap():
                      soup = BeautifulSoup(driver.page_source, 'html.parser')
                      link_source = soup.find('tbody').find_all('a')
                      table = soup.find_all('tbody')
                      table_branch =table[0].find_all('tr')
                      all_links = []
                      global counter
                      nonlocal j,J
                      ws = wb["Alabma_Trademark"]
                      for link in link_source:
                          href = link.get('href')
                          all_links.append('http://arc-sos.state.al.us' + href)
                      matching = [valid_links for valid_links in all_links if "http://arc-sos.state.al.us/cgi/tmdetail.mbr/detail?trade" in valid_links]
                      for table_branch_row in table_branch[:-1]:
                         table_branch_cell = table_branch_row.find_all('td')
                         Applicant_name = table_branch_cell[1].get_text()
                         Mark_Description = table_branch_cell[2].get_text()
                         ws.cell(j, 3).value = Applicant_name
                         ws.cell(j, 4).value = Mark_Description
                         ws.cell(j, 1).value = keywords
                         ws.cell(j, 2).value = "Alabma"
                         if keywords.count(keywords) == Mark_Description.count(Mark_Description) == 1:
                             counter = counter + 1
                         j = j + 1
                      for Data, s_llink in enumerate(matching):
                         source = requests.get(s_llink).text
                         soup2 = BeautifulSoup(source, 'html.parser')
                         data_source = soup2.find_all('td', {'class': 'aiSosDetailValue'})
                         i = 5
                         for data in data_source:
                             datafill = data.get_text()
                             ws.cell(J, i).value = datafill
                             i = i + 1
                         J = J + 1
                      return scrap
                  if __name__ == '__main__':
                     scrap = scrap()
                  while True:
                      try:
                          WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, "Next >>"))).click()
                          scrap()
                      except TimeoutException:
                          break

              global counter
              ws1.cell(rownumber, 2).value = "Albma Trademark"
              ws1.cell(rownumber, 3).value = counter
              counter = False
              counter = 0
              driver.quit()
    wb.save(file_name)
    print("Execution time1 = {0:.5f}".format(time.time() - start))
def Alabma_Bussiness():
    file_name = "Data for 50 State.xlsx"
    wb = openpyxl.load_workbook(file_name)
    ws1 = wb.worksheets[0]
    N = range(2, ws1.max_row + 1)
    j = 2
    J = 2
    for rownumber in N:
        keywords = ws1.cell(rownumber, 1).value
        if keywords is not None:
            driver = webdriver.Chrome(chrome_options=options)
            driver.get('http://arc-sos.state.al.us/CGI/CORPNAME.MBR/INPUT')
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.visibility_of_any_elements_located((By.TAG_NAME, "input")))
            driver.find_element_by_xpath('//*[@id="block-sos-content"]/div/div/div[1]/form/div[1]/input').send_keys(
                keywords)
            driver.find_element_by_xpath('//*[@id="block-sos-content"]/div/div/div[1]/form/div[6]/input').click()
            element1 = wait.until(EC.visibility_of_any_elements_located((By.TAG_NAME, "input")))
            frame = driver.find_elements_by_tag_name('tbody')
            if not frame:
                ws1.cell(rownumber, 5).value = 0
            else:
                def scrap():
                    soup = BeautifulSoup(driver.page_source, 'html.parser')
                    link_source = soup.find('tbody').find_all('a')
                    table = soup.find_all('tbody')
                    table_branch = table[0].find_all('tr')
                    all_links = []
                    newlist = []
                    global counter
                    nonlocal j, J
                    ws = wb["Alabma_Bussiness"]
                    for link in link_source:
                        href = link.get('href')
                        all_links.append('http://arc-sos.state.al.us' + href)
                    matching = [valid_links for valid_links in all_links if
                                "http://arc-sos.state.al.us/cgi/corpdetail.mbr/detail?corp" in valid_links]
                    for exact_matching in matching:
                        if exact_matching not in newlist:
                            newlist.append(exact_matching)
                    for table_branch_row in table_branch[:-1]:
                        table_branch_cell = table_branch_row.find_all('td')
                        Entity_Name = table_branch_cell[1].get_text()
                        City = table_branch_cell[2].get_text()
                        ws.cell(j, 3).value = Entity_Name
                        ws.cell(j, 4).value = City
                        ws.cell(j, 1).value = keywords
                        ws.cell(j, 2).value = "Alabma"
                        j = j + 1
                        if keywords.count(keywords) == Entity_Name.count(Entity_Name) == 1:
                            counter = counter + 1
                    for Data, s_llink in enumerate(newlist):
                        source = requests.get(s_llink).text
                        soup2 = BeautifulSoup(source, 'html.parser')
                        data_source = soup2.find_all('td', {'class': 'aiSosDetailValue'})
                        i = 5
                        for data in data_source:
                            datafill = data.get_text()
                            ws.cell(J, i).value = datafill
                            i = i + 1
                        J = J + 1
                    return scrap
                if __name__ == '__main__':
                    scrap = scrap()
                while True:
                    try:
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, "Next >>"))).click()
                        scrap()
                    except TimeoutException:
                        break
            global counter
            ws1.cell(rownumber, 4).value = "Albma Bussiness"
            ws1.cell(rownumber, 5).value = counter
            counter = False
            counter = 0
            driver.quit()
    wb.save(file_name)
    print("Execution time2 = {0:.5f}".format(time.time() - start))
def California():
    file_name = "Data for 50 State.xlsx"
    wb = openpyxl.load_workbook(file_name)
    ws1 = wb.worksheets[0]
    N = range(2, ws1.max_row + 1)
    j = 2
    J = 2
    for rownumber in N:
        keywords = ws1.cell(rownumber, 1).value
        if keywords is not None:
            driver = webdriver.Chrome(chrome_options=options)
            driver.get('https://businesssearch.sos.ca.gov/')
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.visibility_of_any_elements_located((By.TAG_NAME, "input")))
            driver.find_element_by_id('CorpNameOpt').click()
            driver.find_element_by_id('SearchCriteria').send_keys(keywords)
            driver.find_element_by_tag_name('button').click()
            frame = driver.find_elements_by_link_text("1")

            if not frame:
                ws1.cell(rownumber, 7).value = 0
            else:
                def scrap():
                    soup = BeautifulSoup(driver.page_source, 'html.parser')
                    link_source = soup.find('tbody')
                    table_branch = link_source.find_all('tr')
                    global counter
                    nonlocal j, J
                    next_button = driver.find_element_by_css_selector(
                        "#enitityTable_paginate > ul > li.paginate_button.active > a")
                    next_button_text =next_button.text
                    ws = wb["California"]
                    total_id=[]
                    for table_branch_row in table_branch:
                        driver.find_element_by_link_text(next_button_text).click()
                        table_branch_cell = table_branch_row.find_all('td')
                        Entity_number_data = table_branch_cell[0].get_text()
                        Entity_name_data = table_branch_cell[3].get_text()
                        ws.cell(j, 1).value = keywords
                        ws.cell(j, 2).value = "California"
                        Entity_number =Entity_number_data.replace(" ","").strip()
                        ID = Entity_number.replace('C', '0')
                        ws.cell(j, 4).value = Entity_number
                        Entity_name = Entity_name_data.replace('View details for entity number',"").replace(ID ,"").strip()
                        ws.cell(j, 3).value = Entity_name
                        if keywords.count(keywords) == Entity_name.count(Entity_name) == 1:
                            counter = counter + 1
                        j = j + 1
                        total_id.append('btnDetail-' + ID)
                    for click_id, Id_value in enumerate(total_id):
                        driver.find_element_by_id(Id_value).click()
                        WebDriverWait(driver, 10).until(EC.visibility_of_any_elements_located((By.TAG_NAME, 'div')))
                        soup2 = BeautifulSoup(driver.page_source, 'html.parser')
                        table = soup2.find_all('div', {'id': "maincontent"})
                        for data in table:
                            datasource = data.find_all('div', {'class': "col-sm-8 col-xs-6"})
                            i = 5
                            for data in datasource:
                                datafill = data.get_text()
                                filterdata = datafill.replace("Entity", "").replace("Agent", "").replace("City",
                                                                                                         "").replace(
                                    "State", "").replace("Zip", "").replace("Address", "").replace(",", "").replace(
                                    "Mailing", "")
                                ws.cell(J, i).value = filterdata
                                i = i + 1
                            J = J + 1
                        driver.back()
                        WebDriverWait(driver,10).until(EC.visibility_of_element_located((By.ID, "enitityTable_info")))
                        driver.find_element_by_link_text(next_button_text).click()
                    return scrap
                if __name__ == '__main__':
                    scrap = scrap()
                while True:
                    check1 = driver.find_element_by_id('enitityTable_info').text
                    driver.find_element_by_xpath('//*[@id="enitityTable_next"]/a').click()
                    check2 = driver.find_element_by_id('enitityTable_info').text
                    if check2 == check1:
                        break
                    else:
                        scrap()
            global counter
            ws1.cell(rownumber, 6).value = "California"
            ws1.cell(rownumber, 7).value = counter
            counter = False
            counter = 0
            driver.quit()
    wb.save(file_name)
    print("Execution time3 = {0:.5f}".format(time.time() - start))
def Colorado_Bussiness():
    file_name = "Data for 50 State.xlsx"
    wb = openpyxl.load_workbook(file_name)
    ws1 = wb.worksheets[0]
    N = range(2, ws1.max_row + 1)
    j = 2
    J = 2
    for rownumber in N:
        keywords = ws1.cell(rownumber, 1).value
        if keywords is not None:
            driver = webdriver.Chrome(chrome_options=options)
            if driver.get('https://www.sos.state.co.us/biz/BusinessEntityCriteriaExt.do'):
               WebDriverWait(driver, 10)
            else:driver.get('https://www.sos.state.co.us/biz/BusinessEntityCriteriaExt.do')
            wait = WebDriverWait(driver, 10)
            wait.until(EC.visibility_of_any_elements_located((By.TAG_NAME, "input")))
            driver.find_element_by_name('searchName').send_keys(keywords)
            driver.find_element_by_xpath(
                '//*[@id="application"]/table/tbody/tr/td[2]/table/tbody/tr[3]/td/form/table[2]/tbody/tr/td[1]/input').click()
            frame = driver.find_elements_by_tag_name('caption')
            if not frame:
                ws1.cell(rownumber, 9).value = 0
            else:
                def scrap():
                    soup = BeautifulSoup(driver.page_source, 'html.parser')
                    table = driver.find_element_by_id('box')
                    table_branch = table.find_elements_by_tag_name('tr')
                    global counter
                    nonlocal j, J
                    for table_branch_row in table_branch[1:]:
                        table_branch_cell = table_branch_row.find_element_by_tag_name('a')
                        Entity_number = table_branch_cell.text
                        ws = wb["Colorado_Bussiness"]
                        ws.cell(j, 1).value = keywords
                        ws.cell(j, 2).value = "Colorado"
                        ws.cell(j, 3).value = Entity_number
                        if keywords.count(keywords) == Entity_number.count(Entity_number) == 1:
                            counter = counter + 1
                        j = j + 1
                        data_source = table_branch_row.find_elements_by_tag_name('td')
                        i= 4
                        for data in data_source[3:]:
                            datafill = data.text
                            ws.cell(J, i).value =  datafill
                            i = i + 1
                        J = J + 1
                    return scrap
                if __name__ == '__main__':
                    scrap = scrap()
                while True:
                    try:
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Next"))).click()
                        scrap()
                    except TimeoutException:
                        break
            global counter
            ws1.cell(rownumber, 8).value = "Colorado_Bussiness"
            ws1.cell(rownumber, 9).value = counter
            counter = False
            counter = 0
            driver.quit()
    wb.save(file_name)
    print("Execution time4 = {0:.5f}".format(time.time() - start))
def Colorado_Trademark():
    file_name = "Data for 50 State.xlsx"
    wb = openpyxl.load_workbook(file_name)
    ws1 = wb.worksheets[0]
    N = range(2, ws1.max_row + 1)
    j = 2
    J = 2
    for rownumber in N:
        keywords = ws1.cell(rownumber, 1).value
        if keywords is not None:
            driver = webdriver.Chrome(chrome_options=options)
            if driver.get('https://www.sos.state.co.us/biz/AdvancedTrademarkSearchCriteria.do'):
               WebDriverWait(driver, 10)
            else:
                driver.get('https://www.sos.state.co.us/biz/AdvancedTrademarkSearchCriteria.do')
            wait = WebDriverWait(driver, 10)
            element = wait.until(EC.visibility_of_any_elements_located((By.TAG_NAME, "input")))
            driver.find_element_by_name('trademark').send_keys(keywords)
            driver.find_element_by_name('cmd').click()
            frame = driver.find_elements_by_tag_name('caption')
            if not frame:
                ws1.cell(rownumber, 10).value = 0
            else:
                def scrap():
                    soup = BeautifulSoup(driver.page_source, 'html.parser')
                    link_source = soup.find('div', {'id': 'box'}).find_all('a')
                    all_links = []
                    global counter
                    nonlocal j, J
                    for link in link_source:
                        href = link.get('href')
                        all_links.append("https://www.sos.state.co.us/biz/" + href)
                    matching = [valid_links for valid_links in all_links if
                                 "https://www.sos.state.co.us/biz/T" in valid_links]
                    for Data, s_llink in enumerate(matching):
                        browser = webdriver.Chrome(chrome_options=options)
                        if browser.get(s_llink):
                           WebDriverWait(browser, 10)
                        else:
                            browser.get(s_llink)
                        WebDriverWait(browser, 10)
                        table_branch = browser.find_element_by_name('TradeMarkDetailForm')
                        table_branch_row = table_branch.find_elements_by_tag_name('tr')
                        i= 4
                        table_branch_cell = table_branch_row[1].find_elements_by_tag_name('td')
                        ws = wb["Colorado_Trademark"]
                        ws.cell(j, 1).value = keywords
                        ws.cell(j, 2).value = "Colorado"
                        for data2 in table_branch_cell[2:3]:
                            Trademark = data2.text
                            ws.cell(j,3).value = Trademark
                            if keywords.count(keywords) == Trademark.count(Trademark) == 1:
                                counter = counter + 1
                            j = j + 1
                        for data in table_branch_cell[:2:-2]:
                            data_fill = data.text
                            ws.cell(J, i).value = data_fill
                            i = i + 1
                        J = J + 1
                        browser.close()
                    return scrap
                if __name__ == '__main__':
                    scrap = scrap()
                while True:
                    try:
                        WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.PARTIAL_LINK_TEXT, "Next"))).click()
                        scrap()
                    except TimeoutException:
                        break
            global counter
            ws1.cell(rownumber, 10).value = "Colorado_Trademark"
            ws1.cell(rownumber, 11).value = counter
            counter = False
            counter = 0
            driver.quit()
    wb.save(file_name)
    print("Execution time5 = {0:.5f}".format(time.time() - start))
if __name__ == '__main__':
    try:Alabma_Trademark()
    except:pass
    try:Alabma_Bussiness()
    except:pass
    try:California()
    except:pass
    try:Colorado_Bussiness()
    except:pass
    try:Colorado_Trademark()
    except:pass
    print("Execution time = {0:.5f}".format(time.time() - start))

